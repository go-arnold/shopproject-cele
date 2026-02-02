from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from django.contrib.auth.models import Group
from shop.models import Vente
from utils.decorators import admin_required
from django.core.exceptions import PermissionDenied
from django.templatetags.static import static
import os


@admin_required
def export_ventes_excel(request):
    user = request.user
    rev = Group.objects.get(name="revendeur").user_set.all()
    muk = Group.objects.get(name="mukubwa").user_set.all()

    if user in rev and user not in muk:
        raise PermissionDenied()

    if user.groups.filter(name="mukubwa").exists():
        ventes = Vente.objects.select_related("produit", "utilisateur")
    else:
        ventes = Vente.objects.filter(utilisateur=user).select_related(
            "produit", "utilisateur"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventes"

    # logo_path = os.path.join("static-img", "logo-white.png") 
    logo_path = request.build_absolute_uri(static("static-img/logo-white.png"))

    if os.path.exists(logo_path):
        logo = XLImage(logo_path)
        logo.height = 80  
        logo.width = 120
        ws.add_image(logo, "C1")  

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4B0082")
    alignment_center = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["Produit", "Utilisateur", "Prix (USD)", "Méthode", "Date"]

   
    header_row = 5
    ws.append([""] * len(headers)) 
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border

    fill_even = PatternFill("solid", fgColor="E6E6FA")  
    fill_odd = PatternFill("solid", fgColor="F5F5F5")  

    for row_num, v in enumerate(ventes, start=header_row + 1):
        prix = float(v.price_final or v.produit_prix)
        values = [
            v.produit_nom,
            v.utilisateur.username,
            prix,
            v.method,
            v.date_achat.strftime("%Y-%m-%d %H:%M"),
        ]

        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = val
            cell.fill = fill_even if row_num % 2 == 0 else fill_odd
            cell.border = thin_border
            if col_num == 3:  
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
        ws.column_dimensions[column].width = max_length + 5

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ventes.xlsx"'
    wb.save(response)
    return response
